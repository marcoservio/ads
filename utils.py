from seleniumbase import Driver
from selenium.webdriver.common.by import By
from main import SCROLL_AMOUNT, PUBLI_CLASS
import openpyxl
import time
import os


def start_navegador():
    DIR = "C:\\Users\Marco Sérvio\\AppData\\Local\\Google\\Chrome\\User Data"

    driver = Driver(
        uc=True,
        headed=False,
        undetectable=True,
        undetected=True,
        headless=False,
        user_data_dir=DIR,
        agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    )

    driver.get("https://www.instagram.com/")
    time.sleep(1)

    return driver


def fechar_aba_anuncio(driver):
    janela_inicial = driver.current_window_handle

    for aba in driver.window_handles:
        if aba != janela_inicial:
            driver.switch_to.window(aba)
            driver.close()

    driver.switch_to.window(janela_inicial)


def carregar_urls_existente(worksheet):
    return {
        cell.hyperlink.target
        for row in worksheet.iter_rows(min_row=1, max_col=3, max_row=worksheet.max_row)
        for cell in row
        if cell.hyperlink
    }


def salvar_excel(link, quantidade_like, descricao, nome_do_arquivo):
    if os.path.exists(nome_do_arquivo):
        workbook = openpyxl.load_workbook(nome_do_arquivo)
        sheet = workbook.active
        proxima_linha = sheet.max_row + 1
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        proxima_linha = 1

    urls_existentes = carregar_urls_existente(sheet)

    if link not in urls_existentes:
        sheet.cell(row=proxima_linha, column=1, value=quantidade_like)
        sheet.cell(row=proxima_linha, column=2, value=descricao)
        sheet.cell(row=proxima_linha, column=3, value=link)
        sheet.cell(row=proxima_linha, column=3).hyperlink = link

    workbook.save(nome_do_arquivo)


def scrollar(driver):
    driver.execute_script(f"window.scrollBy(0, {SCROLL_AMOUNT});")
    divs_publicacao = driver.find_elements(By.CSS_SELECTOR, PUBLI_CLASS)
    time.sleep(1)
    return divs_publicacao
